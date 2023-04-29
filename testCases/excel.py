import openpyxl
import pytest
list=[]


path = "D:\\Credence Python Projects by Tushar Sir\\OrangeHRM\\TestData\\EmployeeList.xlsx"

book =openpyxl.load_workbook("D:\\Credence Python Projects by Tushar Sir\\OrangeHRM\\TestData\\EmployeeList.xlsx")
book =openpyxl.load_workbook(path)
sheet =book.active
row = sheet.max_row
print(row)
for r in range(2,row+1):
    firstname = sheet.cell(r, 2).value
    middlename = sheet.cell(r, 3).value
    lastname = sheet.cell(r, 4).value
    empid = sheet.cell(r, 5).value
    username = sheet.cell(r, 6).value
    password = sheet.cell(r, 7).value
    confrimpassword = sheet.cell(r, 7).value

    tuple =(firstname, middlename, lastname, empid, username, password, confrimpassword)
    list.append(tuple)

print(list)